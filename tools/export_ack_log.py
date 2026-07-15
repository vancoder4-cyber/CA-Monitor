# -*- coding: utf-8 -*-
"""把『确认留痕库』(data/ack_log.json)导成一份 Excel。

留痕库是只追加、永不删的审计底账:每一次人工『确认』(含改值)都会记一条。
本脚本随时可跑,产出可离线归档/发风控的留痕表。

用法:
    # 从线上仓库拉(默认,需 GH_TOKEN,与机器人同一套环境变量):
    python tools/export_ack_log.py [输出.xlsx]

    # 从本地 json 文件导(离线/调试):
    python tools/export_ack_log.py --from data/ack_log.json [输出.xlsx]

环境变量(线上模式):
    GH_TOKEN   细粒度 PAT(对本仓库 Contents 有读权限即可)
    GH_REPO    默认 vancoder4-cyber/CA-Monitor
    GH_BRANCH  默认 main
"""
import os
import sys
import json
import base64
import datetime as dt

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API = "https://api.github.com"
LOG_PATH = "data/ack_log.json"

COLUMNS = [
    ("序号", 6),
    ("确认时间(北京)", 18),
    ("标的", 8),
    ("事件类型", 10),
    ("关键日(除息/生效)", 16),
    ("确认值", 12),
    ("上一个值", 12),
    ("核对来源(点开核对)", 34),
    ("备注(核对了什么)", 26),
    ("确认人", 14),
    ("确认人 open_id", 26),
]
ETYPE_CN = {"dividend": "分红", "split": "拆股", "filing": "并购/公告"}


def fetch_log_from_github():
    token = os.environ.get("GH_TOKEN", "").strip()
    repo = os.environ.get("GH_REPO", "vancoder4-cyber/CA-Monitor").strip()
    branch = os.environ.get("GH_BRANCH", "main").strip()
    if not token:
        sys.exit("✗ 没有 GH_TOKEN。请设置环境变量,或用 --from <本地json> 离线导出。")
    url = f"{API}/repos/{repo}/contents/{LOG_PATH}?ref={branch}"
    r = requests.get(url, headers={
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28"}, timeout=20)
    if r.status_code == 404:
        return []
    if r.status_code != 200:
        sys.exit(f"✗ 拉取失败 HTTP {r.status_code}: {r.text[:200]}")
    return json.loads(base64.b64decode(r.json()["content"]).decode("utf-8"))


def load_log(args):
    if "--from" in args:
        i = args.index("--from")
        path = args[i + 1]
        del args[i:i + 2]
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    return fetch_log_from_github()


def bj(iso):
    """ISO(带 +08:00)→ '2026-06-25 21:46'。"""
    s = iso or ""
    return f"{s[:10]} {s[11:16]}".strip() if len(s) >= 16 else s


def build(rows, out):
    # 时间倒序(最新在前)
    rows = sorted(rows, key=lambda e: e.get("at_utc") or e.get("at_bj") or "", reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "确认留痕"

    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    link_font = Font(color="1155CC", underline="single")
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")

    for c, (name, w) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, name)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = w

    for i, e in enumerate(rows, 1):
        r = i + 1
        vals = [
            i,
            bj(e.get("at_bj") or e.get("at_utc")),
            e.get("ticker", ""),
            ETYPE_CN.get(e.get("etype"), e.get("etype") or ""),
            e.get("date", "") or "",
            "" if e.get("value") in (None, "") else e.get("value"),
            "" if e.get("prev_value") in (None, "") else e.get("prev_value"),
            e.get("source", "") or "",
            e.get("note", "") or "",
            e.get("by_name", "") or "",
            e.get("by", "") or "",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = wrap
            cell.border = border
            if c == 8 and v:  # 核对来源做成超链接
                cell.hyperlink = v
                cell.value = "打开核对页"
                cell.font = link_font
        # 改了值的行整行淡黄,提醒『这是一次值变更』
        if e.get("prev_value") not in (None, "", e.get("value")):
            for c in range(1, len(COLUMNS) + 1):
                ws.cell(r, c).fill = PatternFill("solid", fgColor="FFF8E1")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(rows)+1, 1)}"
    ws.row_dimensions[1].height = 30

    # 说明页
    doc = wb.create_sheet("说明")
    notes = [
        ["确认留痕库 · 导出说明", ""],
        ["", ""],
        ["这是什么", "每次在群里 @机器人 发『确认』都会往留痕库追加一条(只追加、永不删)。本表是它的快照。"],
        ["谁 / 何时", "确认人(显示名,取不到则 open_id)+ 确认时间(北京)。"],
        ["关键日", "这一列是该事件的『除息日(分红)/ 生效日(拆股)』,不是登记日、也不是派发日。异常就是按这个日期归组和确认的。"],
        ["改了什么", "确认值 = 本次确认的值;上一个值 = 改之前的值。两者不同的行整行标黄 = 一次值变更,重点复核。"],
        ["核对来源", "系统按标的+事件类型自动带的最权威来源:并购/退市→SEC 原文;分红→公司 IR / Nasdaq;拆股→公司 IR / SEC 8-K。点开自行核对。"],
        ["备注", "确认人在命令末尾写的核对说明(如『已比对公司 8-K』)。"],
        ["怎么刷新", "重跑 `python tools/export_ack_log.py`(需 GH_TOKEN)即拉最新;或群里发『留痕』看最近几条。"],
        ["", ""],
        ["确认命令格式", "确认 代码 [正确值] [日期] [备注]"],
        ["示例", "确认 KLAC 2.3 2026-05-18 已比对公司8-K"],
    ]
    doc.column_dimensions["A"].width = 16
    doc.column_dimensions["B"].width = 92
    for r, (a, b) in enumerate(notes, 1):
        doc.cell(r, 1, a).font = Font(bold=(r == 1 or (b and a and r > 2)), size=(13 if r == 1 else 11))
        doc.cell(r, 2, b).alignment = Alignment(wrap_text=True, vertical="center")

    wb.save(out)
    print(f"✓ 已导出 {len(rows)} 条留痕 → {out}")


def main():
    args = sys.argv[1:]
    rows = load_log(args)
    out = args[0] if args else f"确认留痕_{dt.date.today().isoformat()}.xlsx"
    build(rows, out)


if __name__ == "__main__":
    main()
